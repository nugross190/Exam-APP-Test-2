"""Excel/CSV parser for HADIR Exam App.

Implements spec section 2:
  - parse_students(xi_path, x_path) -> ParseResult[StudentRow]
  - parse_schedule(schedule_path)   -> ParseResult[ScheduleEntry]
  - derive_class_subjects(entries)  -> dict[str, set[str]]

Contract per spec: never raise on bad rows. Collect warnings, continue.

NOTE: parse_schedule now accepts a pre-parsed CSV (long/tidy format)
produced by database/schedule_parser.py. Expected columns:
  kelas, subject, date (YYYY-MM-DD), time_start (HH:MM), time_end (HH:MM)
"""
from __future__ import annotations

import csv
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, time
from typing import Iterable

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Result containers
# ---------------------------------------------------------------------------

@dataclass
class ParseResult:
    data: list = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class StudentRow:
    nisn: str
    nis: str
    name: str
    gender: str
    kelas: str
    flags: list[str] = field(default_factory=list)


@dataclass
class ScheduleEntry:
    kelas: str
    subject: str
    date: date
    time_start: time
    time_end: time


# ---------------------------------------------------------------------------
# 2.1  parse_students
# ---------------------------------------------------------------------------

_NISN_IDX   = 3
_NIS_IDX    = 4
_NAME_IDX   = 5
_GENDER_IDX = 6
_KELAS_IDX  = 7


def parse_students(xi_path: str, x_path: str) -> ParseResult:
    """Parse both grade XI and grade X student rosters.

    Dedup state is shared across both files: a NISN that appears in XI
    and again in X is flagged 'nisn_dup' on its second occurrence.
    """
    result = ParseResult()
    nisn_seen: set[str] = set()
    nis_seen: set[str] = set()

    files = [
        (xi_path, "daftar_hadir_kelas_XI_updated"),
        (x_path,  "daftar_hadir_kelas_X_updated"),
    ]

    for path, sheet_name in files:
        wb = load_workbook(path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            result.warnings.append(
                f"sheet_missing: {sheet_name!r} not in {path}"
            )
            wb.close()
            continue
        ws = wb[sheet_name]

        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            first = row[0]
            if not isinstance(first, int) or isinstance(first, bool):
                continue
            if len(row) <= _KELAS_IDX:
                result.warnings.append(
                    f"short_row: row {first} in {sheet_name} has only "
                    f"{len(row)} cols"
                )
                continue

            nisn   = _str_or_empty(row[_NISN_IDX])
            nis    = _str_or_empty(row[_NIS_IDX])
            name   = _str_or_empty(row[_NAME_IDX])
            gender = _str_or_empty(row[_GENDER_IDX])
            kelas  = _str_or_empty(row[_KELAS_IDX])

            flags: list[str] = []
            if len(nisn) != 10:
                flags.append("nisn_invalid")
            if nisn in nisn_seen:
                flags.append("nisn_dup")
            if nis in nis_seen:
                flags.append("nis_dup")
            nisn_seen.add(nisn)
            nis_seen.add(nis)

            result.data.append(StudentRow(
                nisn=nisn, nis=nis, name=name,
                gender=gender, kelas=kelas, flags=flags,
            ))
        wb.close()

    return result


def _str_or_empty(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


# ---------------------------------------------------------------------------
# 2.2  parse_schedule  (CSV version)
# ---------------------------------------------------------------------------

_REQUIRED_COLS = {"kelas", "subject", "date", "time_start", "time_end"}


def parse_schedule(schedule_path: str) -> ParseResult:
    """Parse the pre-processed schedule CSV (long/tidy format).

    Expected columns: kelas, subject, date (YYYY-MM-DD),
    time_start (HH:MM), time_end (HH:MM).

    Produced by database/schedule_parser.py from the original xlsx grid.
    Never raises — bad rows go to warnings.
    """
    result = ParseResult()

    try:
        f = open(schedule_path, newline="", encoding="utf-8")
    except OSError as e:
        result.warnings.append(f"file_open_error: {e}")
        return result

    with f:
        reader = csv.DictReader(f)

        # Validate headers exist
        if not reader.fieldnames:
            result.warnings.append("empty_file: no headers found")
            return result

        missing = _REQUIRED_COLS - set(reader.fieldnames)
        if missing:
            result.warnings.append(
                f"missing_columns: {sorted(missing)} not in CSV headers"
            )
            return result

        for lineno, row in enumerate(reader, start=2):  # 1-indexed, row 1 = header
            kelas   = row["kelas"].strip()
            subject = row["subject"].strip()
            date_s  = row["date"].strip()
            ts_s    = row["time_start"].strip()
            te_s    = row["time_end"].strip()

            if not kelas or not subject:
                result.warnings.append(
                    f"line {lineno}: empty kelas or subject, skipping"
                )
                continue

            try:
                d = _parse_iso_date(date_s)
            except ValueError as e:
                result.warnings.append(f"line {lineno}: bad date {date_s!r}: {e}")
                continue

            try:
                ts = _parse_hhmm(ts_s)
                te = _parse_hhmm(te_s)
            except ValueError as e:
                result.warnings.append(f"line {lineno}: bad time {ts_s!r}/{te_s!r}: {e}")
                continue

            result.data.append(ScheduleEntry(
                kelas=kelas, subject=subject,
                date=d, time_start=ts, time_end=te,
            ))

    return result


def _parse_iso_date(s: str) -> date:
    """'2025-12-08' -> date(2025, 12, 8). Also handles '2025-12-8'."""
    parts = s.split("-")
    if len(parts) != 3:
        raise ValueError(f"expected YYYY-MM-DD, got {s!r}")
    return date(int(parts[0]), int(parts[1]), int(parts[2]))


def _parse_hhmm(s: str) -> time:
    """'07:30' -> time(7, 30)."""
    parts = s.split(":")
    if len(parts) != 2:
        raise ValueError(f"expected HH:MM, got {s!r}")
    return time(int(parts[0]), int(parts[1]))


# ---------------------------------------------------------------------------
# 2.3  derive_class_subjects
# ---------------------------------------------------------------------------

def derive_class_subjects(entries: Iterable[ScheduleEntry]) -> dict[str, set[str]]:
    out: dict[str, set[str]] = defaultdict(set)
    for e in entries:
        out[e.kelas].add(e.subject)
    return dict(out)